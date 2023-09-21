import openpyxl

def convert_mm_to_cm(value):
    return value / 10

# 读取两个xlsx文件
template_wb = openpyxl.load_workbook('template.xlsx')
rasson_wb = openpyxl.load_workbook('mlds-list.xlsx')

template_ws = template_wb["SKUs"]
rasson_ws = rasson_wb.active

# 获取work.xlsx中的必要列索引
sku_col = 1
packing_size_col = 2
weight_col = 3

# 从work.xlsx中创建SKU到详情的映射
sku_details = {}

for row in range(2, rasson_ws.max_row + 1):
    sku = rasson_ws.cell(row=row, column=sku_col).value
    packing_size_raw = rasson_ws.cell(row=row, column=packing_size_col).value
    weight = rasson_ws.cell(row=row, column=weight_col).value

    if sku not in sku_details:
        sku_details[sku] = {"weight": [], "length": [], "width": [], "height": []}

    try:
        length, width, height = map(lambda x: convert_mm_to_cm(float(x)), packing_size_raw.split('*'))

        sku_details[sku]["weight"].append(weight)
        sku_details[sku]["length"].append(length)
        sku_details[sku]["width"].append(width)
        sku_details[sku]["height"].append(height)
    except Exception as e:
        print(f"在第{row}行遇到了问题: {e}")

# 更新"SKUs"工作表中的数据
for row in range(2, template_ws.max_row + 1):
    sku = template_ws.cell(row=row, column=1).value
    details = sku_details.get(sku, None)
    if details:
        for idx, (weight, length, width, height) in enumerate(zip(details["weight"], details["length"], details["width"], details["height"]), start=1):
            template_ws.cell(row=row, column=1 + (idx-1)*5 + 1).value = weight
            template_ws.cell(row=row, column=1 + (idx-1)*5 + 2).value = length
            template_ws.cell(row=row, column=1 + (idx-1)*5 + 3).value = width
            template_ws.cell(row=row, column=1 + (idx-1)*5 + 4).value = height
            template_ws.cell(row=row, column=1 + (idx-1)*5 + 5).value = 1

# 更新"Total"工作表的数据
total_ws = template_wb["Total"]
current_row = 2
for sku, details in sku_details.items():
    total_weight = sum(details["weight"])
    max_length = max(details["length"])
    max_width = max(details["width"])
    total_height = sum(details["height"])
    total_ws.cell(row=current_row, column=1).value = sku
    total_ws.cell(row=current_row, column=2).value = total_weight
    total_ws.cell(row=current_row, column=3).value = max_length
    total_ws.cell(row=current_row, column=4).value = max_width
    total_ws.cell(row=current_row, column=5).value = total_height
    current_row += 1

template_wb.save('processed_template1.xlsx')
